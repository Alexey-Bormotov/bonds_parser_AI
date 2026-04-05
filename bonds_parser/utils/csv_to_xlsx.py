"""
Утилита для конвертации CSV файла в формат XLSX.
Предназначена для автоматического запуска после завершения работы spider.
"""
import csv
import logging
import sys
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)


def convert_csv_to_xlsx(
    csv_path: str,
    xlsx_path: Optional[str] = None,
    delimiter: str = ";",
    encoding: str = "utf-8",
) -> str:
    """
    Конвертирует CSV файл в XLSX.

    Args:
        csv_path: Путь к исходному CSV файлу
        xlsx_path: Путь для сохранения XLSX файла (если None, генерируется автоматически)
        delimiter: Разделитель в CSV (по умолчанию точка с запятой)
        encoding: Кодировка CSV файла

    Returns:
        Путь к созданному XLSX файлу

    Raises:
        FileNotFoundError: Если CSV файл не существует
        ValueError: Если CSV файл пуст или повреждён
    """
    csv_file = Path(csv_path)
    if not csv_file.exists():
        raise FileNotFoundError(f"CSV файл не найден: {csv_path}")

    if xlsx_path is None:
        xlsx_path = str(csv_file.with_suffix(".xlsx"))

    logger.info(f"Конвертация CSV -> XLSX: {csv_path} -> {xlsx_path}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Облигации"

    try:
        with open(csv_path, "r", encoding=encoding, newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            for row_idx, row in enumerate(reader):
                for col_idx, value in enumerate(row):
                    ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)

        # Определяем индекс колонки "Рейтинг" по заголовку (первая строка)
        rating_col_index = None
        first_row = [cell.value for cell in ws[1]]
        for idx, header in enumerate(first_row, start=1):
            if header == "Рейтинг":
                rating_col_index = idx
                break
        
        # Если не нашли по заголовку, используем фиксированный индекс 8 (колонка H, 7-я по счёту с 0)
        if rating_col_index is None:
            rating_col_index = 8  # Колонка H соответствует индексу 8 (A=1, B=2, ... H=8)
            logger.warning("Заголовок 'Рейтинг' не найден, используется фиксированная колонка 8")

        # Создаём заливку бледно-зелёного цвета
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

        # Применяем заливку к строкам, где рейтинг равен "AAA"
        for row in ws.iter_rows(min_row=2):  # Пропускаем заголовок
            rating_cell = row[rating_col_index - 1]  # индекс с 1
            if rating_cell.value == "AAA":
                for cell in row:
                    cell.fill = green_fill

        wb.save(xlsx_path)
        logger.info(f"XLSX файл успешно создан: {xlsx_path}")
        return xlsx_path
    except Exception as e:
        logger.error(f"Ошибка при конвертации CSV в XLSX: {e}")
        raise


def main():
    """Точка входа для запуска утилиты из командной строки."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Конвертирует CSV файл (разделитель ';') в XLSX"
    )
    parser.add_argument(
        "csv_path", help="Путь к CSV файлу (например, data/bonds_output.csv)"
    )
    parser.add_argument(
        "-o",
        "--output",
        help="Путь для сохранения XLSX файла (по умолчанию: тот же путь с расширением .xlsx)",
    )
    parser.add_argument(
        "-d", "--delimiter", default=";", help="Разделитель в CSV (по умолчанию ';')"
    )
    parser.add_argument(
        "-e", "--encoding", default="utf-8", help="Кодировка CSV файла"
    )
    parser.add_argument(
        "-v", "--verbose", action="store_true", help="Вывод подробных сообщений"
    )

    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    )

    try:
        convert_csv_to_xlsx(
            csv_path=args.csv_path,
            xlsx_path=args.output,
            delimiter=args.delimiter,
            encoding=args.encoding,
        )
        sys.exit(0)
    except Exception as e:
        logger.error(f"Конвертация не удалась: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()